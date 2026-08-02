from __future__ import annotations

from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_store import DistributionStore

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "distribution_demo.db"
SAMPLE_XLSX = DATA_DIR / "remitos_demo.xlsx"
EXPORT_DIR = BASE_DIR / "exports"

BANNERS = {
    "DEPORTE": ["LOCAL CENTRO", "LOCAL NORTE", "E-COMMERCE", "ALMACÉN DEPORTE"],
    "MODA": ["MODA STORE", "LOCAL SUR", "ALMACÉN MODA"],
    "JB": ["LOCAL JB", "ALMACÉN JB"],
}

COLORS = {
    "navy": "#122033", "blue": "#2457D6", "bg": "#F4F6FA", "card": "#FFFFFF",
    "text": "#182230", "muted": "#667085", "border": "#DDE3EA", "green": "#137A55",
    "orange": "#B54708", "red": "#B42318", "selected": "#EAF0FF",
}


class DistributionManager(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Distribution Manager · Portfolio Demo")
        self.geometry("1280x790")
        self.minsize(1080, 680)
        self.configure(bg=COLORS["bg"])
        self.store = DistributionStore(DB_PATH)
        self.remito_actual: str | None = None
        self.articulo_actual: str | None = None
        self.stock_actual: dict[str, int] = {}
        self.entries: dict[tuple[str, str], ttk.Entry] = {}
        self.row_total_labels: dict[str, ttk.Label] = {}
        self._configure_style()
        self._build_ui()
        self._load_demo_if_empty()
        self.refresh_all()

    def _configure_style(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background=COLORS["bg"])
        style.configure("Card.TFrame", background=COLORS["card"], relief="flat")
        style.configure("TLabel", background=COLORS["bg"], foreground=COLORS["text"], font=("Segoe UI", 10))
        style.configure("Card.TLabel", background=COLORS["card"], foreground=COLORS["text"])
        style.configure("Title.TLabel", font=("Segoe UI Semibold", 21), foreground="white", background=COLORS["navy"])
        style.configure("Subtitle.TLabel", font=("Segoe UI", 10), foreground="#C9D4E5", background=COLORS["navy"])
        style.configure("Metric.TLabel", font=("Segoe UI Semibold", 22), background=COLORS["card"], foreground=COLORS["text"])
        style.configure("MetricName.TLabel", font=("Segoe UI", 9), background=COLORS["card"], foreground=COLORS["muted"])
        style.configure("Primary.TButton", font=("Segoe UI Semibold", 10), padding=(14, 9), background=COLORS["blue"], foreground="white")
        style.map("Primary.TButton", background=[("active", "#1947BC")])
        style.configure("Secondary.TButton", font=("Segoe UI", 10), padding=(12, 8), background="white", foreground=COLORS["text"], bordercolor=COLORS["border"])
        style.configure("TNotebook", background=COLORS["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", font=("Segoe UI Semibold", 10), padding=(18, 10), background="#E8ECF2", foreground=COLORS["muted"])
        style.map("TNotebook.Tab", background=[("selected", "white")], foreground=[("selected", COLORS["blue"])])
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=31, background="white", fieldbackground="white", foreground=COLORS["text"], bordercolor=COLORS["border"])
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 9), background="#EEF2F7", foreground=COLORS["text"], relief="flat", padding=7)
        style.map("Treeview", background=[("selected", COLORS["selected"])], foreground=[("selected", COLORS["text"])])
        style.configure("TEntry", padding=6)
        style.configure("TCombobox", padding=6)

    def _build_ui(self) -> None:
        header = tk.Frame(self, bg=COLORS["navy"], height=104)
        header.pack(fill="x")
        header.pack_propagate(False)
        title_wrap = ttk.Frame(header, style="TFrame")
        title_wrap.configure(style="Header.TFrame")
        ttk.Style(self).configure("Header.TFrame", background=COLORS["navy"])
        title_wrap.pack(side="left", padx=24, pady=16)
        ttk.Label(title_wrap, text="Distribution Manager", style="Title.TLabel").pack(anchor="w")
        ttk.Label(title_wrap, text="Asignación y control de mercadería · Demo pública con datos ficticios", style="Subtitle.TLabel").pack(anchor="w", pady=(3, 0))

        actions = tk.Frame(header, bg=COLORS["navy"])
        actions.pack(side="right", padx=24)
        ttk.Button(actions, text="Importar Excel", command=self.load_excel, style="Primary.TButton").pack(side="left", padx=4)
        ttk.Button(actions, text="Restablecer demo", command=self.reset_demo, style="Secondary.TButton").pack(side="left", padx=4)
        ttk.Button(actions, text="Exportar", command=self.export_report, style="Secondary.TButton").pack(side="left", padx=4)

        self.metrics_frame = ttk.Frame(self, padding=(18, 14), style="TFrame")
        self.metrics_frame.pack(fill="x")
        self.metric_labels: dict[str, ttk.Label] = {}
        for index, (key, label) in enumerate((("remitos", "Remitos"), ("articulos", "Artículos"), ("unidades", "Unidades"), ("asignados", "Artículos asignados"))):
            card = ttk.Frame(self.metrics_frame, padding=(18, 12), style="Card.TFrame")
            card.grid(row=0, column=index, sticky="ew", padx=5)
            self.metrics_frame.columnconfigure(index, weight=1)
            value = ttk.Label(card, text="0", style="Metric.TLabel")
            value.pack(anchor="w")
            ttk.Label(card, text=label.upper(), style="MetricName.TLabel").pack(anchor="w")
            self.metric_labels[key] = value

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=22, pady=(0, 18))
        self.summary_tab = ttk.Frame(self.notebook, padding=14, style="Card.TFrame")
        self.assignment_tab = ttk.Frame(self.notebook, padding=14, style="Card.TFrame")
        self.distribution_tab = ttk.Frame(self.notebook, padding=14, style="Card.TFrame")
        self.reports_tab = ttk.Frame(self.notebook, padding=14, style="Card.TFrame")
        self.notebook.add(self.summary_tab, text="Resumen")
        self.notebook.add(self.assignment_tab, text="Asignación")
        self.notebook.add(self.distribution_tab, text="Distribución")
        self.notebook.add(self.reports_tab, text="Reportes")
        self._build_summary()
        self._build_assignment()
        self._build_distribution()
        self._build_reports()

    def _tree(self, parent: ttk.Frame, columns: tuple[str, ...], widths: tuple[int, ...]) -> ttk.Treeview:
        wrapper = ttk.Frame(parent, style="Card.TFrame")
        wrapper.pack(fill="both", expand=True)
        tree = ttk.Treeview(wrapper, columns=columns, show="headings")
        ybar = ttk.Scrollbar(wrapper, orient="vertical", command=tree.yview)
        xbar = ttk.Scrollbar(wrapper, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ybar.set, xscrollcommand=xbar.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ybar.grid(row=0, column=1, sticky="ns")
        xbar.grid(row=1, column=0, sticky="ew")
        wrapper.rowconfigure(0, weight=1)
        wrapper.columnconfigure(0, weight=1)
        for column, width in zip(columns, widths):
            tree.heading(column, text=column.replace("_", " ").upper())
            tree.column(column, width=width, minwidth=70, anchor="center")
        return tree

    def _build_summary(self) -> None:
        top = ttk.Frame(self.summary_tab, style="Card.TFrame")
        top.pack(fill="x", pady=(0, 10))
        ttk.Label(top, text="Remitos y artículos", style="Card.TLabel", font=("Segoe UI Semibold", 14)).pack(side="left")
        ttk.Label(top, text="Doble clic para comenzar la distribución", style="Card.TLabel", foreground=COLORS["muted"]).pack(side="right")
        columns = ("remito", "articulo", "descripcion", "marca", "total", "estado", "empresa")
        self.summary_tree = self._tree(self.summary_tab, columns, (105, 135, 310, 100, 85, 120, 140))
        self.summary_tree.bind("<Double-1>", self.open_selected_article)
        self.summary_tree.tag_configure("asignado", foreground=COLORS["green"])
        self.summary_tree.tag_configure("no_trabajado", foreground=COLORS["orange"])

    def _build_assignment(self) -> None:
        top = ttk.Frame(self.assignment_tab, style="Card.TFrame")
        top.pack(fill="x", pady=(0, 12))
        self.article_label = ttk.Label(top, text="Seleccioná un artículo desde Resumen", style="Card.TLabel", font=("Segoe UI Semibold", 14))
        self.article_label.pack(side="left")
        controls = ttk.Frame(top, style="Card.TFrame")
        controls.pack(side="right")
        ttk.Label(controls, text="Banner", style="Card.TLabel").pack(side="left", padx=(0, 6))
        self.banner_var = tk.StringVar(value=next(iter(BANNERS)))
        self.banner_combo = ttk.Combobox(controls, textvariable=self.banner_var, values=list(BANNERS), state="readonly", width=13)
        self.banner_combo.pack(side="left")
        self.banner_combo.bind("<<ComboboxSelected>>", lambda _event: self.build_assignment_grid())

        grid_shell = ttk.Frame(self.assignment_tab, style="Card.TFrame")
        grid_shell.pack(fill="both", expand=True)
        self.assignment_canvas = tk.Canvas(grid_shell, bg="white", highlightthickness=0)
        ybar = ttk.Scrollbar(grid_shell, orient="vertical", command=self.assignment_canvas.yview)
        xbar = ttk.Scrollbar(grid_shell, orient="horizontal", command=self.assignment_canvas.xview)
        self.assignment_grid = ttk.Frame(self.assignment_canvas, padding=4, style="Card.TFrame")
        self.assignment_window = self.assignment_canvas.create_window((0, 0), window=self.assignment_grid, anchor="nw")
        self.assignment_canvas.configure(yscrollcommand=ybar.set, xscrollcommand=xbar.set)
        self.assignment_canvas.grid(row=0, column=0, sticky="nsew")
        ybar.grid(row=0, column=1, sticky="ns")
        xbar.grid(row=1, column=0, sticky="ew")
        grid_shell.rowconfigure(0, weight=1)
        grid_shell.columnconfigure(0, weight=1)
        self.assignment_grid.bind("<Configure>", lambda _e: self.assignment_canvas.configure(scrollregion=self.assignment_canvas.bbox("all")))

        bottom = ttk.Frame(self.assignment_tab, style="Card.TFrame")
        bottom.pack(fill="x", pady=(12, 0))
        self.remainder_label = ttk.Label(bottom, text="", style="Card.TLabel", foreground=COLORS["muted"])
        self.remainder_label.pack(side="left")
        ttk.Button(bottom, text="Guardar asignación", command=self.save_assignment, style="Primary.TButton").pack(side="right")

    def _build_distribution(self) -> None:
        ttk.Label(self.distribution_tab, text="Distribución consolidada", style="Card.TLabel", font=("Segoe UI Semibold", 14)).pack(anchor="w", pady=(0, 10))
        columns = ("remito", "articulo", "banner", "local", "talle", "cantidad")
        self.distribution_tree = self._tree(self.distribution_tab, columns, (105, 135, 110, 230, 85, 95))

    def _build_reports(self) -> None:
        filters = ttk.Frame(self.reports_tab, style="Card.TFrame")
        filters.pack(fill="x", pady=(0, 12))
        ttk.Label(filters, text="Artículo", style="Card.TLabel").pack(side="left")
        self.article_filter = tk.StringVar()
        ttk.Entry(filters, textvariable=self.article_filter, width=18).pack(side="left", padx=(6, 16))
        ttk.Label(filters, text="Local", style="Card.TLabel").pack(side="left")
        self.local_filter = tk.StringVar()
        ttk.Entry(filters, textvariable=self.local_filter, width=20).pack(side="left", padx=(6, 12))
        ttk.Button(filters, text="Buscar", command=self.refresh_reports, style="Primary.TButton").pack(side="left", padx=4)
        ttk.Button(filters, text="Limpiar", command=self.clear_filters, style="Secondary.TButton").pack(side="left", padx=4)
        columns = ("remito", "articulo", "descripcion", "marca", "banner", "local", "talle", "cantidad", "estado")
        self.reports_tree = self._tree(self.reports_tab, columns, (100, 125, 250, 90, 100, 210, 75, 90, 110))

    def _load_demo_if_empty(self) -> None:
        if not self.store.summary() and SAMPLE_XLSX.exists():
            self.store.import_dataframe(pd.read_excel(SAMPLE_XLSX))

    def load_excel(self) -> None:
        path = filedialog.askopenfilename(title="Seleccionar remito", filetypes=[("Archivo Excel", "*.xlsx")])
        if not path:
            return
        try:
            count = self.store.import_dataframe(pd.read_excel(path))
            self.refresh_all()
            messagebox.showinfo("Importación completa", f"Se importaron o actualizaron {count} artículos.")
        except Exception as exc:
            messagebox.showerror("No se pudo importar", str(exc))

    def reset_demo(self) -> None:
        if not messagebox.askyesno("Restablecer demo", "Se eliminarán las asignaciones actuales y se recargarán los datos ficticios. ¿Continuar?"):
            return
        try:
            self.store.reset()
            self.store.import_dataframe(pd.read_excel(SAMPLE_XLSX))
            self.remito_actual = self.articulo_actual = None
            self.stock_actual = {}
            self.article_label.config(text="Seleccioná un artículo desde Resumen")
            self.build_assignment_grid()
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def refresh_all(self) -> None:
        self.refresh_metrics()
        self.refresh_summary()
        self.refresh_distribution()
        self.refresh_reports()

    def refresh_metrics(self) -> None:
        metrics = self.store.metrics()
        for key, label in self.metric_labels.items():
            label.config(text=f"{metrics.get(key, 0):,}".replace(",", "."))

    def refresh_summary(self) -> None:
        self.summary_tree.delete(*self.summary_tree.get_children())
        for row in self.store.summary():
            self.summary_tree.insert("", "end", values=tuple(row), tags=(row[5],))

    def open_selected_article(self, _event=None) -> None:
        selected = self.summary_tree.selection()
        if not selected:
            return
        values = self.summary_tree.item(selected[0], "values")
        self.remito_actual, self.articulo_actual = str(values[0]), str(values[1])
        self.notebook.select(self.assignment_tab)
        self.build_assignment_grid()

    def build_assignment_grid(self) -> None:
        for widget in self.assignment_grid.winfo_children():
            widget.destroy()
        self.entries.clear()
        self.row_total_labels.clear()
        if not self.remito_actual or not self.articulo_actual:
            self.remainder_label.config(text="")
            return
        try:
            description, self.stock_actual, previous = self.store.article_detail(self.remito_actual, self.articulo_actual)
        except LookupError as exc:
            messagebox.showerror("Artículo no disponible", str(exc))
            return
        sizes = list(self.stock_actual)
        previous_map = {(row[1], row[2]): int(row[3]) for row in previous if row[0] == self.banner_var.get()}
        locations = [location for location in BANNERS[self.banner_var.get()] if "ALMACÉN" not in location]
        self.article_label.config(text=f"{self.remito_actual}  ·  {self.articulo_actual}  ·  {description}")

        headers = ["DESTINO", *sizes, "TOTAL"]
        for col, text in enumerate(headers):
            ttk.Label(self.assignment_grid, text=text, style="Card.TLabel", font=("Segoe UI Semibold", 9)).grid(row=0, column=col, padx=7, pady=(4, 8), sticky="ew")
        for row_index, location in enumerate(locations, start=1):
            ttk.Label(self.assignment_grid, text=location, style="Card.TLabel").grid(row=row_index, column=0, padx=7, pady=5, sticky="w")
            for col_index, size in enumerate(sizes, start=1):
                entry = ttk.Entry(self.assignment_grid, width=9, justify="center")
                value = previous_map.get((location, size), 0)
                if value:
                    entry.insert(0, str(value))
                entry.grid(row=row_index, column=col_index, padx=5, pady=5)
                entry.bind("<KeyRelease>", lambda _event: self.update_remainder())
                self.entries[(location, size)] = entry
            total = ttk.Label(self.assignment_grid, text="0", style="Card.TLabel", foreground=COLORS["muted"])
            total.grid(row=row_index, column=len(sizes) + 1, padx=7)
            self.row_total_labels[location] = total
        stock_row = len(locations) + 2
        ttk.Label(self.assignment_grid, text="STOCK DISPONIBLE", style="Card.TLabel", font=("Segoe UI Semibold", 9)).grid(row=stock_row, column=0, padx=7, pady=(14, 4), sticky="w")
        for col_index, size in enumerate(sizes, start=1):
            ttk.Label(self.assignment_grid, text=str(self.stock_actual[size]), style="Card.TLabel", foreground=COLORS["blue"], font=("Segoe UI Semibold", 10)).grid(row=stock_row, column=col_index)
        self.update_remainder()

    def _read_entries(self) -> list[tuple[str, str, int]]:
        result: list[tuple[str, str, int]] = []
        for (location, size), entry in self.entries.items():
            text = entry.get().strip()
            if not text:
                continue
            if not text.isdigit():
                raise ValueError(f"Ingresá un entero válido en {location}, talle {size}.")
            result.append((location, size, int(text)))
        return result

    def update_remainder(self) -> None:
        if not self.stock_actual:
            self.remainder_label.config(text="")
            return
        try:
            values = self._read_entries()
        except ValueError:
            self.remainder_label.config(text="Revisá los valores ingresados.", foreground=COLORS["red"])
            return
        totals = {size: 0 for size in self.stock_actual}
        for _location, size, quantity in values:
            totals[size] += quantity
        location_totals = {location: 0 for location in self.row_total_labels}
        for location, _size, quantity in values:
            if location in location_totals:
                location_totals[location] += quantity
        for location, label in self.row_total_labels.items():
            label.config(text=str(location_totals[location]))
        parts = [f"{size}: {self.stock_actual[size] - totals[size]}" for size in self.stock_actual]
        has_negative = any(self.stock_actual[size] - totals[size] < 0 for size in self.stock_actual)
        self.remainder_label.config(text="Restante por talle  ·  " + "   |   ".join(parts) + "   ·   El remanente se envía automáticamente al almacén.", foreground=COLORS["red"] if has_negative else COLORS["muted"])

    def save_assignment(self) -> None:
        if not self.remito_actual or not self.articulo_actual:
            messagebox.showwarning("Sin artículo", "Elegí un artículo desde la pestaña Resumen.")
            return
        try:
            values = self._read_entries()
            banner = self.banner_var.get()
            warehouse = next(location for location in BANNERS[banner] if "ALMACÉN" in location)
            self.store.save_assignment(self.remito_actual, self.articulo_actual, banner, values, warehouse)
            self.refresh_all()
            self.build_assignment_grid()
            messagebox.showinfo("Asignación guardada", "La distribución se guardó correctamente. El remanente fue enviado al almacén.")
        except Exception as exc:
            messagebox.showerror("No se pudo guardar", str(exc))

    def refresh_distribution(self) -> None:
        self.distribution_tree.delete(*self.distribution_tree.get_children())
        for row in self.store.distribution():
            self.distribution_tree.insert("", "end", values=tuple(row))

    def refresh_reports(self) -> None:
        self.reports_tree.delete(*self.reports_tree.get_children())
        df = self.store.report(self.article_filter.get(), self.local_filter.get())
        for row in df.itertuples(index=False, name=None):
            self.reports_tree.insert("", "end", values=row)

    def clear_filters(self) -> None:
        self.article_filter.set("")
        self.local_filter.set("")
        self.refresh_reports()

    def export_report(self) -> None:
        EXPORT_DIR.mkdir(exist_ok=True)
        path = filedialog.asksaveasfilename(
            title="Exportar distribución", initialdir=EXPORT_DIR,
            initialfile="distribution_report.xlsx", defaultextension=".xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            df = self.store.report()
            df.to_excel(path, index=False, sheet_name="Distribución")
            workbook = load_workbook(path)
            sheet = workbook["Distribución"]
            header_fill = PatternFill("solid", fgColor="122033")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in range(1, sheet.max_column + 1):
                values = [str(sheet.cell(row=row, column=column).value or "") for row in range(1, sheet.max_row + 1)]
                sheet.column_dimensions[get_column_letter(column)].width = min(max(len(value) for value in values) + 2, 40)
            workbook.save(path)
            messagebox.showinfo("Reporte exportado", f"Archivo generado correctamente:\n{path}")
        except Exception as exc:
            messagebox.showerror("No se pudo exportar", str(exc))


if __name__ == "__main__":
    DistributionManager().mainloop()
